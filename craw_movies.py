# -*- coding:utf-8 -*-

"""

@Author: Truman_Lv
@file: Scraw_movies.py
@time: 2017/8/4 21:53
"""

import sys
import importlib
import time
import requests
import urllib.parse
import urllib.request
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

importlib.reload(sys)


# Some User Agents

hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
       {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def movie_spider(movie_tag):
    page_num = 0

    movie_list = []

    try_times = 0

    while (1):

        url = 'http://www.douban.com/tag/' + urllib.parse.quote(movie_tag) + '/movie?start=' + str(page_num * 15)

        time.sleep(np.random.rand() * 5)

        try:

            req = requests.get(url, headers=hds[page_num % len(hds)])

            req_text = req.text

        except (urllib.request.HTTPError, urllib.request.URLError) as e:

            print(e)

            continue

        soup = BeautifulSoup(req_text, "lxml")

        list_soup = soup.find('div', {'class': 'mod movie-list'})

        try_times += 1

        if list_soup is None and try_times < 200:

            continue

        elif list_soup is None or len(list_soup) <= 1:

            break  # Break when no information got after 200 times requesting

        for movie_info in list_soup.findAll('dd'):

            title = movie_info.find('a', {'class': 'title'}).string.strip()

            desc = movie_info.find('div', {'class': 'desc'}).string.strip()

            desc_list = desc.split('/')

            movie_url = movie_info.find('a', {'class': 'title'}).get('href')

            try:

                area_info = '地区： ' + desc_list[0]

            except:

                area_info = '地区： 暂无'

            try:

                rating = movie_info.find('span', {'class': 'rating_nums'}).string.strip()

            except:

                rating = '0.0'

            try:

                people_num = get_people_info(movie_url)

                people_num = people_num.strip('人评价')

            except:

                people_num = '0'

            movie_list.append([title, rating, people_num, area_info])

            try_times = 0  # set 0 when got valid information

        page_num += 1

        print('Downloading Information From Page %d' % page_num)

    return movie_list


def get_people_info(url):

    global req_text

    try:

        req = requests.get(url, headers=hds[np.random.randint(0, len(hds))])
        req_text = req.text

    except (urllib.request.HTTPError, urllib.request.URLError) as e:

        print(e)

    soup = BeautifulSoup(req_text, "lxml")
    people_num = soup.find('div', {'class': 'rating_sum'}).findAll('span')[0].string.strip()
    return people_num


def do_spider(movie_tag_lists):

    movie_lists = []

    for movie_tag in movie_tag_lists:
        movie_list = movie_spider(movie_tag)

        movie_list = sorted(movie_list, key=lambda x: x[1], reverse=True)

        movie_lists.append(movie_list)

    return movie_lists


def print_book_lists_excel(movie_lists, movie_tag_lists):
    wb = Workbook(optimized_write=True)

    ws = []

    for i in range(len(movie_tag_lists)):
        ws.append(wb.create_sheet(title=movie_tag_lists[i]))

    for i in range(len(movie_tag_lists)):

        ws[i].append(['序号', '电影名', '评分', '评价人数', '地区'])

        count = 1

        for bl in movie_lists[i]:
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3]])

            count += 1

    save_path = 'movie_list'

    for i in range(len(movie_tag_lists)):
        save_path += ('-' + movie_tag_lists[i])

    save_path += '.xlsx'

    wb.save(save_path)


if __name__ == '__main__':

    # movie_tag_lists = ['李安', '是枝裕和']

    # movie_tag_lists = ['恋童癖', '恋童']

    movie_tag_lists = ['ellenpage']

    movie_lists = do_spider(movie_tag_lists)

    print_book_lists_excel(movie_lists, movie_tag_lists) 